#!/usr/bin/env python3
"""
Reconcile current cities.json with INE 2026 data (diccionario26.xlsx).
Adds new municipalities, renames updated ones, and preserves existing image URLs.
"""

import json
import os
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(ROOT, "diccionario26.xlsx")
CITIES_PATH = os.path.join(ROOT, "src", "data", "cities.json")
REPORT_PATH = os.path.join(ROOT, "scripts", "reconciliation_report.txt")

NO_FLAG = "https://raw.githubusercontent.com/ByMykel/spanish-cities/refs/heads/main/no_flag.svg"
NO_COAT = "https://raw.githubusercontent.com/ByMykel/spanish-cities/refs/heads/main/no_coat.svg"


def parse_xlsx(path):
    """Parse diccionario26.xlsx and return dict of code -> {code, name, code_autonomy, code_province}."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    municipalities = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        codauto, cpro, cmun, dc, nombre = row
        if codauto is None or nombre is None:
            continue
        code_autonomy = str(codauto).zfill(2)
        code_province = str(cpro).zfill(2)
        cmun_str = str(cmun).zfill(3)
        dc_str = str(dc)
        code = f"{code_province}{cmun_str}{dc_str}"
        municipalities[code] = {
            "code": code,
            "name": nombre.strip(),
            "code_autonomy": code_autonomy,
            "code_province": code_province,
        }
    wb.close()
    return municipalities


def load_cities(path):
    """Load current cities.json and return dict of code -> city object."""
    with open(path, "r", encoding="utf-8") as f:
        cities_list = json.load(f)
    return {c["code"]: c for c in cities_list}


def reconcile():
    ine_data = parse_xlsx(XLSX_PATH)
    current_cities = load_cities(CITIES_PATH)

    report_lines = []
    report_lines.append(f"INE 2026 Reconciliation Report")
    report_lines.append(f"{'=' * 50}")
    report_lines.append(f"INE municipalities: {len(ine_data)}")
    report_lines.append(f"Current cities.json: {len(current_cities)}")
    report_lines.append("")

    # Find additions, removals, renames
    ine_codes = set(ine_data.keys())
    current_codes = set(current_cities.keys())

    added_codes = ine_codes - current_codes
    removed_codes = current_codes - ine_codes
    common_codes = ine_codes & current_codes

    renames = []
    for code in sorted(common_codes):
        old_name = current_cities[code]["name"]
        new_name = ine_data[code]["name"]
        if old_name != new_name:
            renames.append((code, old_name, new_name))

    # Report additions
    report_lines.append(f"ADDITIONS ({len(added_codes)}):")
    report_lines.append("-" * 40)
    for code in sorted(added_codes):
        m = ine_data[code]
        report_lines.append(f"  {code} - {m['name']} (prov: {m['code_province']}, auto: {m['code_autonomy']})")
    report_lines.append("")

    # Report removals
    report_lines.append(f"REMOVALS ({len(removed_codes)}):")
    report_lines.append("-" * 40)
    for code in sorted(removed_codes):
        c = current_cities[code]
        report_lines.append(f"  {code} - {c['name']} (prov: {c['code_province']}, auto: {c['code_autonomy']})")
    report_lines.append("")

    # Report renames
    report_lines.append(f"RENAMES ({len(renames)}):")
    report_lines.append("-" * 40)
    for code, old_name, new_name in renames:
        report_lines.append(f"  {code}: '{old_name}' -> '{new_name}'")
    report_lines.append("")

    # Build updated cities list
    updated_cities = []
    for code in sorted(ine_data.keys()):
        ine_entry = ine_data[code]
        if code in current_cities:
            # Keep existing city, update name from INE
            city = current_cities[code].copy()
            city["name"] = ine_entry["name"]
            # Also update autonomy/province codes in case of changes
            city["code_autonomy"] = ine_entry["code_autonomy"]
            city["code_province"] = ine_entry["code_province"]
            updated_cities.append(city)
        else:
            # New city - add with placeholder images
            updated_cities.append({
                "code": code,
                "name": ine_entry["name"],
                "code_autonomy": ine_entry["code_autonomy"],
                "code_province": ine_entry["code_province"],
                "flag": NO_FLAG,
                "coat_of_arms": NO_COAT,
            })

    # Write updated cities.json
    with open(CITIES_PATH, "w", encoding="utf-8") as f:
        json.dump(updated_cities, f, ensure_ascii=False, indent=2)

    report_lines.append(f"RESULT:")
    report_lines.append(f"  Updated cities.json: {len(updated_cities)} cities")
    report_lines.append(f"  Added: {len(added_codes)}, Removed: {len(removed_codes)}, Renamed: {len(renames)}")

    # Write report
    report_text = "\n".join(report_lines)
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write(report_text)

    print(report_text)
    print(f"\nReport saved to: {REPORT_PATH}")
    print(f"Cities saved to: {CITIES_PATH}")


if __name__ == "__main__":
    reconcile()
